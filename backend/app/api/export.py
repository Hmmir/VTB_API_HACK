"""Export endpoints."""
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime, timedelta
from io import BytesIO
from typing import Optional
from app.database import get_db
from app.api.dependencies import get_current_user
from app.models.user import User
from app.models.transaction import Transaction, TransactionType
from app.models.account import Account
from sqlalchemy import func

# PDF Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

# Excel Generation  
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ODS Generation (for Мой Офис)
try:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.style import Style, TextProperties, ParagraphProperties, TableColumnProperties
    from odf.text import P
    from odf.table import Table as ODSTable, TableColumn, TableRow, TableCell
    ODS_AVAILABLE = True
except ImportError:
    ODS_AVAILABLE = False

router = APIRouter()


def generate_transaction_report_csv(
    user: User,
    db: Session,
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None
) -> BytesIO:
    """Generate CSV report of transactions."""
    
    # Get user's accounts
    account_ids = [acc.id for acc in db.query(Account).join(Account.bank_connection).filter(
        Account.bank_connection.has(user_id=user.id)
    ).all()]
    
    # Build query
    query = db.query(Transaction).filter(Transaction.account_id.in_(account_ids))
    
    if from_date:
        query = query.filter(Transaction.transaction_date >= from_date)
    if to_date:
        query = query.filter(Transaction.transaction_date <= to_date)
    
    transactions = query.order_by(Transaction.transaction_date.desc()).all()
    
    # Generate CSV
    output = BytesIO()
    
    # Header
    header = "Дата,Описание,Категория,Тип,Сумма,Валюта,Счет,Банк\n"
    output.write(header.encode('utf-8-sig'))
    
    # Rows
    for tx in transactions:
        account = db.query(Account).filter(Account.id == tx.account_id).first()
        bank = account.bank_connection.bank_provider.value if account and account.bank_connection else "Unknown"
        
        row = f"{tx.transaction_date.strftime('%Y-%m-%d %H:%M:%S')},"
        row += f'"{tx.description}",'
        row += f'"{tx.category.name if tx.category else "Без категории"}",'
        row += f"{tx.transaction_type.value},"
        row += f"{tx.amount},"
        row += f"{tx.currency},"
        row += f'"{account.account_name if account else "Unknown"}",'
        row += f'"{bank}"\n'
        
        output.write(row.encode('utf-8-sig'))
    
    output.seek(0)
    return output


@router.get("/transactions/csv")
def export_transactions_csv(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export transactions to CSV format."""
    
    # Parse dates
    from_dt = None
    to_dt = None
    
    if from_date:
        try:
            from_dt = datetime.fromisoformat(from_date)
        except:
            pass
    
    if to_date:
        try:
            to_dt = datetime.fromisoformat(to_date)
        except:
            pass
    
    # Default to last 30 days if no dates provided
    if not from_dt:
        from_dt = datetime.utcnow() - timedelta(days=30)
    if not to_dt:
        to_dt = datetime.utcnow()
    
    csv_data = generate_transaction_report_csv(current_user, db, from_dt, to_dt)
    
    filename = f"transactions_{from_dt.strftime('%Y%m%d')}_{to_dt.strftime('%Y%m%d')}.csv"
    
    return StreamingResponse(
        csv_data,
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/analytics/csv")
def export_analytics_csv(
    period_days: int = Query(30, ge=1, le=365),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export analytics summary to CSV."""
    from app.models.category import Category
    
    # Get user's accounts
    account_ids = [acc.id for acc in db.query(Account).join(Account.bank_connection).filter(
        Account.bank_connection.has(user_id=current_user.id)
    ).all()]
    
    if not account_ids:
        return StreamingResponse(
            BytesIO(b"No data available"),
            media_type="text/csv"
        )
    
    # Date range
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=period_days)
    
    # Get expenses by category
    results = db.query(
        Category.name,
        func.sum(Transaction.amount).label("total"),
        func.count(Transaction.id).label("count")
    ).join(
        Transaction, Transaction.category_id == Category.id
    ).filter(
        Transaction.account_id.in_(account_ids),
        Transaction.transaction_type == TransactionType.EXPENSE,
        Transaction.transaction_date >= start_date,
        Transaction.transaction_date <= end_date
    ).group_by(
        Category.id, Category.name
    ).order_by(
        func.sum(Transaction.amount).desc()
    ).all()
    
    # Generate CSV
    output = BytesIO()
    
    header = "Категория,Сумма (₽),Количество транзакций\n"
    output.write(header.encode('utf-8-sig'))
    
    for row in results:
        line = f'"{row.name}",{row.total},{row.count}\n'
        output.write(line.encode('utf-8-sig'))
    
    output.seek(0)
    
    filename = f"analytics_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.csv"
    
    return StreamingResponse(
        output,
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


def generate_transactions_pdf(
    user: User,
    db: Session,
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None
) -> BytesIO:
    """Generate PDF report of transactions with proper Cyrillic support."""
    
    # Register fonts for Cyrillic support
    # Try to register fonts, fallback if not available
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        # Use DejaVuSans which supports Cyrillic
        pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
        font_name = 'DejaVuSans'
        font_name_bold = 'DejaVuSans-Bold'
    except:
        # Fallback to default (will show squares for Cyrillic)
        font_name = 'Helvetica'
        font_name_bold = 'Helvetica-Bold'
    
    # Get user's accounts
    account_ids = [acc.id for acc in db.query(Account).join(Account.bank_connection).filter(
        Account.bank_connection.has(user_id=user.id)
    ).all()]
    
    # Build query
    query = db.query(Transaction).filter(Transaction.account_id.in_(account_ids))
    
    if from_date:
        query = query.filter(Transaction.transaction_date >= from_date)
    if to_date:
        query = query.filter(Transaction.transaction_date <= to_date)
    
    transactions = query.order_by(Transaction.transaction_date.desc()).all()
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles with Cyrillic font
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=font_name_bold,
        fontSize=24,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=12
    )
    
    heading2_style = ParagraphStyle(
        'CustomHeading2',
        parent=styles['Heading2'],
        fontName=font_name_bold,
        fontSize=16
    )
    
    # Title
    title = Paragraph(f"Отчет по транзакциям<br/>FinanceHub", title_style)
    elements.append(title)
    
    # Date range
    period_text = f"Период: {from_date.strftime('%d.%m.%Y') if from_date else '...'} - {to_date.strftime('%d.%m.%Y') if to_date else '...'}"
    elements.append(Paragraph(period_text, normal_style))
    elements.append(Spacer(1, 20))
    
    # Summary statistics
    total_income = sum(tx.amount for tx in transactions if tx.transaction_type == TransactionType.INCOME)
    total_expense = sum(tx.amount for tx in transactions if tx.transaction_type == TransactionType.EXPENSE)
    balance = total_income - total_expense
    
    summary_data = [
        ['Показатель', 'Значение'],
        ['Всего транзакций', str(len(transactions))],
        ['Доходы', f'{total_income:,.2f} ₽'],
        ['Расходы', f'{total_expense:,.2f} ₽'],
        ['Баланс', f'{balance:,.2f} ₽'],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
        ('FONTNAME', (0, 1), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # Transactions table
    if transactions:
        elements.append(Paragraph("Транзакции:", heading2_style))
        elements.append(Spacer(1, 10))
        
        # Table data
        data = [['Дата', 'Описание', 'Категория', 'Тип', 'Сумма']]
        
        # Show ALL transactions, not just first 100
        for tx in transactions:
            data.append([
                tx.transaction_date.strftime('%d.%m.%Y'),
                tx.description[:30] + '...' if len(tx.description) > 30 else tx.description,
                tx.category.name if tx.category else '-',
                'Доход' if tx.transaction_type == TransactionType.INCOME else 'Расход',
                f'{tx.amount:,.2f} ₽'
            ])
        
        trans_table = Table(data, colWidths=[1*inch, 2.5*inch, 1.2*inch, 0.8*inch, 1*inch])
        trans_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        
        elements.append(trans_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_transactions_excel(
    user: User,
    db: Session,
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None
) -> BytesIO:
    """Generate Excel report of transactions."""
    
    # Get user's accounts
    account_ids = [acc.id for acc in db.query(Account).join(Account.bank_connection).filter(
        Account.bank_connection.has(user_id=user.id)
    ).all()]
    
    # Build query
    query = db.query(Transaction).filter(Transaction.account_id.in_(account_ids))
    
    if from_date:
        query = query.filter(Transaction.transaction_date >= from_date)
    if to_date:
        query = query.filter(Transaction.transaction_date <= to_date)
    
    transactions = query.order_by(Transaction.transaction_date.desc()).all()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Транзакции"
    
    # Styles
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Дата', 'Время', 'Описание', 'Категория', 'Тип', 'Сумма (₽)', 'Валюта', 'Счет', 'Банк']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data
    for row_idx, tx in enumerate(transactions, 2):
        account = db.query(Account).filter(Account.id == tx.account_id).first()
        bank = account.bank_connection.bank_provider.value if account and account.bank_connection else "Unknown"
        
        ws.cell(row=row_idx, column=1, value=tx.transaction_date.strftime('%d.%m.%Y'))
        ws.cell(row=row_idx, column=2, value=tx.transaction_date.strftime('%H:%M:%S'))
        ws.cell(row=row_idx, column=3, value=tx.description)
        ws.cell(row=row_idx, column=4, value=tx.category.name if tx.category else 'Без категории')
        ws.cell(row=row_idx, column=5, value='Доход' if tx.transaction_type == TransactionType.INCOME else 'Расход')
        ws.cell(row=row_idx, column=6, value=float(tx.amount))
        ws.cell(row=row_idx, column=7, value=tx.currency)
        ws.cell(row=row_idx, column=8, value=account.account_name if account else 'Unknown')
        ws.cell(row=row_idx, column=9, value=bank)
        
        # Apply border
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = border
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/transactions/pdf")
def export_transactions_pdf(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export transactions to PDF format."""
    
    # Parse dates
    from_dt = None
    to_dt = None
    
    if from_date:
        try:
            from_dt = datetime.fromisoformat(from_date)
        except:
            pass
    
    if to_date:
        try:
            to_dt = datetime.fromisoformat(to_date)
        except:
            pass
    
    # Default to last 30 days if no dates provided
    if not from_dt:
        from_dt = datetime.utcnow() - timedelta(days=30)
    if not to_dt:
        to_dt = datetime.utcnow()
    
    pdf_data = generate_transactions_pdf(current_user, db, from_dt, to_dt)
    
    filename = f"transactions_{from_dt.strftime('%Y%m%d')}_{to_dt.strftime('%Y%m%d')}.pdf"
    
    return StreamingResponse(
        pdf_data,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/transactions/excel")
def export_transactions_excel(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export transactions to Excel format."""
    
    # Parse dates
    from_dt = None
    to_dt = None
    
    if from_date:
        try:
            from_dt = datetime.fromisoformat(from_date)
        except:
            pass
    
    if to_date:
        try:
            to_dt = datetime.fromisoformat(to_date)
        except:
            pass
    
    # Default to last 30 days if no dates provided
    if not from_dt:
        from_dt = datetime.utcnow() - timedelta(days=30)
    if not to_dt:
        to_dt = datetime.utcnow()
    
    excel_data = generate_transactions_excel(current_user, db, from_dt, to_dt)
    
    filename = f"transactions_{from_dt.strftime('%Y%m%d')}_{to_dt.strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        excel_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


def generate_transactions_ods(
    user: User,
    db: Session,
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None
) -> BytesIO:
    """Generate ODS report of transactions (for Мой Офис)."""
    
    if not ODS_AVAILABLE:
        raise ImportError("odfpy library is not installed. Install with: pip install odfpy")
    
    # Get user's accounts
    account_ids = [acc.id for acc in db.query(Account).join(Account.bank_connection).filter(
        Account.bank_connection.has(user_id=user.id)
    ).all()]
    
    # Build query
    query = db.query(Transaction).filter(Transaction.account_id.in_(account_ids))
    
    if from_date:
        query = query.filter(Transaction.transaction_date >= from_date)
    if to_date:
        query = query.filter(Transaction.transaction_date <= to_date)
    
    transactions = query.order_by(Transaction.transaction_date.desc()).all()
    
    # Create ODS document
    doc = OpenDocumentSpreadsheet()
    
    # Create table
    table = ODSTable(name="Транзакции")
    
    # Add header row
    headers = ['Дата', 'Время', 'Описание', 'Категория', 'Тип', 'Сумма (₽)', 'Валюта', 'Счет', 'Банк']
    header_row = TableRow()
    for header in headers:
        cell = TableCell()
        cell.addElement(P(text=header))
        header_row.addElement(cell)
    table.addElement(header_row)
    
    # Add data rows
    for tx in transactions:
        account = db.query(Account).filter(Account.id == tx.account_id).first()
        bank = account.bank_connection.bank_provider.value if account and account.bank_connection else "Unknown"
        
        row = TableRow()
        
        # Date
        cell = TableCell()
        cell.addElement(P(text=tx.transaction_date.strftime('%d.%m.%Y')))
        row.addElement(cell)
        
        # Time
        cell = TableCell()
        cell.addElement(P(text=tx.transaction_date.strftime('%H:%M:%S')))
        row.addElement(cell)
        
        # Description
        cell = TableCell()
        cell.addElement(P(text=tx.description))
        row.addElement(cell)
        
        # Category
        cell = TableCell()
        cell.addElement(P(text=tx.category.name if tx.category else 'Без категории'))
        row.addElement(cell)
        
        # Type
        cell = TableCell()
        cell.addElement(P(text='Доход' if tx.transaction_type == TransactionType.INCOME else 'Расход'))
        row.addElement(cell)
        
        # Amount
        cell = TableCell()
        cell.addElement(P(text=str(float(tx.amount))))
        row.addElement(cell)
        
        # Currency
        cell = TableCell()
        cell.addElement(P(text=tx.currency))
        row.addElement(cell)
        
        # Account
        cell = TableCell()
        cell.addElement(P(text=account.account_name if account else 'Unknown'))
        row.addElement(cell)
        
        # Bank
        cell = TableCell()
        cell.addElement(P(text=bank))
        row.addElement(cell)
        
        table.addElement(row)
    
    doc.spreadsheet.addElement(table)
    
    # Save to buffer
    buffer = BytesIO()
    doc.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/transactions/ods")
def export_transactions_ods(
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export transactions to ODS format (for Мой Офис)."""
    
    if not ODS_AVAILABLE:
        from fastapi import HTTPException
        raise HTTPException(
            status_code=501,
            detail="ODS export is not available. Install odfpy library."
        )
    
    # Parse dates
    from_dt = None
    to_dt = None
    
    if from_date:
        try:
            from_dt = datetime.fromisoformat(from_date)
        except:
            pass
    
    if to_date:
        try:
            to_dt = datetime.fromisoformat(to_date)
        except:
            pass
    
    # Default to last 30 days if no dates provided
    if not from_dt:
        from_dt = datetime.utcnow() - timedelta(days=30)
    if not to_dt:
        to_dt = datetime.utcnow()
    
    ods_data = generate_transactions_ods(current_user, db, from_dt, to_dt)
    
    filename = f"transactions_{from_dt.strftime('%Y%m%d')}_{to_dt.strftime('%Y%m%d')}.ods"
    
    return StreamingResponse(
        ods_data,
        media_type="application/vnd.oasis.opendocument.spreadsheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

